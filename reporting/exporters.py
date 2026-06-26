"""Export a generated schedule to Excel (.xlsx) and PDF, with Turkish labels.

Both functions return an in-memory ``bytes`` buffer so the Streamlit layer can
hand them straight to ``st.download_button`` without touching the filesystem.
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)

from models.assignment import SolveResult
from models.assistant import Assistant
from models.exam import Exam
from reporting.reports import assignments_dataframe, workload_dataframe

_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# Bundled DejaVu fonts so Turkish glyphs (ş, ğ, İ, ı, …) render in the PDF.
# Helvetica's built-in encoding lacks several of these, so we register a
# Unicode TrueType font and fall back to Helvetica only if the files are gone.
_FONT_DIR = Path(__file__).resolve().parent.parent / "assets" / "fonts"


def _register_fonts() -> tuple[str, str]:
    """Register the bundled Unicode font; return (regular, bold) font names."""
    regular = _FONT_DIR / "DejaVuSans.ttf"
    bold = _FONT_DIR / "DejaVuSans-Bold.ttf"
    if regular.exists() and bold.exists():
        if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont("DejaVuSans", str(regular)))
            pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", str(bold)))
        return "DejaVuSans", "DejaVuSans-Bold"
    return "Helvetica", "Helvetica-Bold"


def _write_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=title)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:                       # style the header row
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_idx, column in enumerate(df.columns, start=1):  # autofit-ish
        width = max(len(str(column)),
                    *(len(str(v)) for v in df[column])) + 2 if len(df) else 12
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 50)


def export_excel(result: SolveResult, assistants: List[Assistant],
                 exams: List[Exam]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    _write_sheet(wb, "Atamalar",
                 assignments_dataframe(result, assistants, exams))
    _write_sheet(wb, "İş Yükü", workload_dataframe(result))

    summary = pd.DataFrame({
        "Ölçüt": ["Çözücü durumu", "Bölüm ortalaması", "İş yükü farkı",
                  "Çözüm süresi (sn)"],
        "Değer": [result.status.value, result.department_average,
                  result.spread, result.solve_time_seconds],
    })
    _write_sheet(wb, "Özet", summary)

    if result.warnings:
        _write_sheet(wb, "Adalet Notları",
                     pd.DataFrame({"Uyarı": result.warnings}))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_table(df: pd.DataFrame, body_font: str, bold_font: str) -> Table:
    data = [list(df.columns)] + df.astype(str).values.tolist()
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTNAME", (0, 1), (-1, -1), body_font),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#EEF2FA")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B8C4DC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return table


def export_pdf(result: SolveResult, assistants: List[Assistant],
               exams: List[Exam]) -> bytes:
    body_font, bold_font = _register_fonts()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                            topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    # Point the default paragraph styles at the Unicode font.
    for style_name in ("Normal", "Title", "Heading2"):
        styles[style_name].fontName = (
            bold_font if style_name in ("Title", "Heading2") else body_font)

    story = [Paragraph("Sınav Gözetmenlik Çizelgesi", styles["Title"]),
             Paragraph(
                 f"Durum: {result.status.value} &nbsp;|&nbsp; "
                 f"Bölüm ortalaması: {result.department_average} &nbsp;|&nbsp; "
                 f"İş yükü farkı: {result.spread}", styles["Normal"]),
             Spacer(1, 0.4 * cm)]

    story.append(Paragraph("Atamalar", styles["Heading2"]))
    story.append(_pdf_table(assignments_dataframe(result, assistants, exams),
                            body_font, bold_font))
    story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph("İş Yükü Özeti", styles["Heading2"]))
    story.append(_pdf_table(workload_dataframe(result), body_font, bold_font))

    if result.warnings:
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph("Adalet Notları", styles["Heading2"]))
        for w in result.warnings:
            story.append(Paragraph(f"• {w}", styles["Normal"]))

    if result.diagnostics:
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph("Kısıt İhlali Raporu", styles["Heading2"]))
        for d in result.diagnostics:
            story.append(Paragraph(f"• {d}", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()
